import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth import update_session_auth_hash
from support_app.admins.user_permission_access import  (
    check_requested_user_permissions,
)

from django.conf import settings

from .forms import GrievanceForm, AmaatraForm, get_faq_form, SSMForm, ECTransportationForm, RRTransportationForm, FAQForm, CETRankingForm, PUCUpoloadMarksForm, GroupPermissionForm, DailyReportForm, AdminUserCreationForm, AdminUserEditForm, JEEMain1Form, JEEMain2Form, COMEDKForm
from .models import(
    User, Grievance, Amaatra,  SSM, ECTransportation, RRTransportation, FAQ,  CETRanking, PUCUpoloadMarks, DailyReport, JEEMain1, JEEMain2, COMEDK,
    AmaatraFAQ, AmaatraCategory, SSMFAQ, SSMCategory, PESHospitalFAQ, PESHospitalCategory, PESIMSRFAQ, PESIMSRCategory, PESPublicSchoolFAQ, PESPublicSchoolCategory, PESUIMSRFAQ, PESUIMSRCategory
    )
from django.contrib.auth.models import Group, Permission
from django.contrib import messages
from django.db.models import Q
from fuzzywuzzy import fuzz, process

from django.http import HttpResponse
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta, date, time
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from django.http import JsonResponse
import os
import re

from .image_access_download import generate_signed_url
from google.cloud import storage

# Global model_mapping definition
model_mapping = {
    'amaatra': AmaatraFAQ,
    'ssm': SSMFAQ,
    'pesimsr': PESIMSRFAQ,
    'pesuimsr': PESUIMSRFAQ,
    'pespublicschool': PESPublicSchoolFAQ,
    'peshospital': PESHospitalFAQ,
}
category_model_mapping = {
    'amaatra': AmaatraCategory,
    'ssm': SSMCategory,
    'pesimsr': PESIMSRCategory,
    'pesuimsr': PESUIMSRCategory,
    'pespublicschool': PESPublicSchoolCategory,
    'peshospital': PESHospitalCategory,
}
title_name_mapping = {
    'amaatra': 'The Amaatra Academy',
    'ssm': 'SSM Public School',
    'pesimsr': 'PESIMSR',
    'pesuimsr': 'PESUIMSR',
    'pespublicschool': 'PES Public School',
    'peshospital': 'PES Hospital',
}

def check_faq_requested_user_permissions(user, model_mapping):
    # If the user is a superuser, grant access
    if user.is_superuser:
        return True

    # Construct a list of all permission codenames (e.g., 'add_amaatrafaq', 'add_ssmfaq', etc.)
    required_permissions = [f'add_{model_name}faq' for model_name in model_mapping.keys()]

    # Check if the user has all required permissions
    user_permissions = Permission.objects.filter(
        codename__in=required_permissions,
        group__user=user
    ).values_list('codename', flat=True)

    return all(permission in user_permissions for permission in required_permissions)

def selected_category(request, model_name, category_name):
    model_class = model_mapping.get(model_name)
    if not model_class:
        return render(request, 'error/404.html')
    category_list = AmaatraCategory.objects.all()
    faq_list = search_category(request, AmaatraFAQ, category_name)
    context = {
        'category_list':category_list,
        'category_name': category_name,
        'faq_list': faq_list,
    }
    return render(request, 'base/category.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_faq_requested_user_permissions(user, model_mapping), login_url='access-denied-page')
def all_faq_form(request, model_name):
    model_class = model_mapping.get(model_name)
    if not model_class:
        return render(request, 'error/404.html')
    
    category = category_model_mapping.get(model_name)
    category_list = category.objects.all()

    faq_list = model_class.objects.all()[:10]
    form_class = get_faq_form(model_class)
    form = form_class()

    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            faq = form.save()
            category = form.cleaned_data['category']
            faq.category.set(category)
            faq.save()
            return redirect('add-all-faq', model_name)
    
    title = title_name_mapping.get(model_name.lower())
    context = {
        'title':title,
        'form': form,
        'model_name':model_name,
        'faq_list': faq_list,
        'category_list':category_list,
    }
    return render(request, 'web-forms/all-faq-form.html', context)

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_faq_requested_user_permissions(user, model_mapping), login_url='access-denied-page')
def edit_all_faq(request, model_name, pk):
    model_class = model_mapping.get(model_name)
    if not model_class:
        return render(request, 'error/404.html')

    faq_list = model_class.objects.all()[:10]
    form_class = get_faq_form(model_class)
    edit_faq = model_class.objects.get(pk=pk)
    form = form_class(instance=edit_faq)
    
    if request.method == 'POST':
        form = form_class(request.POST, instance=edit_faq)
        if form.is_valid():
            faq = form.save()
            category = form.cleaned_data['category']
            faq.category.set(category)
            faq.save()
            return redirect('add-all-faq', model_name)
    
    title = title_name_mapping.get(model_name.lower())
    context={
        'title':title,
        'form':form,
        'model_name':model_name,
        'faq_list':faq_list,
    }
    return render(request, 'web-forms/all-faq-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_faq_requested_user_permissions(user, model_mapping), login_url='access-denied-page')
def view_all_faq (request, model_name): 
    q = request.GET.get('search') if request.GET.get('search') != None else ''   
    # Fetch all FAQs
    model_class = model_mapping.get(model_name)
    if not model_class:
        return render(request, 'error/404.html')
    all_faqs = model_class.objects.all()
    faq_list = []

    if q:
        for faq in all_faqs:
            if fuzz.partial_ratio(q.lower(), faq.question.lower()) > 70 or \
                fuzz.partial_ratio(q.lower(), faq.answer.lower()) > 70 or \
                any(fuzz.partial_ratio(q.lower(), category.category.lower()) > 70 for category in faq.category.all()):
                faq_list.append(faq)
        faq_list = sorted(faq_list, key=lambda faq: fuzz.partial_ratio(q.lower(), faq.question.lower()), reverse=True)
    else:
        # faq_list = all_faqs.filter(Q(question__icontains=q) | Q(answer__icontains=q) | Q(category__category__icontains=q)).order_by('category__category')
        faq_list = all_faqs


    page_number = request.GET.get('page')
    list_par_page = 25
    faq_list= all_paginator(faq_list, page_number, list_par_page)
    
    title = title_name_mapping.get(model_name.lower())
    
    context = {
        'title':title,
        'model_name':model_name,
        'faq_list':faq_list,
    }
    return render(request, 'users/all-faq.html', context)

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_faq_requested_user_permissions(user, model_mapping), login_url='access-denied-page')
def delete_selected_faq(request, model_name, pk):
    model_class = model_mapping.get(model_name)
    if not model_class:
        return render(request, 'error/404.html')
    delete_faq = get_object_or_404(model_class, pk=pk)
    delete_faq.delete()
    return redirect('selected-faq-list', model_name)



def hn_campus_call_back(request):
    return render(request, 'call-back/hn-index.html')

def ec_campus_call_back(request):
    return render(request, 'call-back/ec-index.html')

def rr_campus_call_back(request):
    return render(request, 'call-back/rr-index.html')


# Function to create a sanitized filename
def create_unique_filename(f):   
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    extension = os.path.splitext(f.name)[1]  # Get the file extension
    new_filename = f"{timestamp}{extension}"
    return new_filename


def comedk_form (request):
    if request.method == 'POST':
        form = COMEDKForm(request.POST, request.FILES)
        if form.is_valid():
            comedk = form.save(commit=False)
            upload_marks = form.cleaned_data.get('upload_marks')
            phone_number = form.cleaned_data.get('phone_number')
            if upload_marks:
                filename = create_unique_filename(upload_marks)
                comedk.upload_marks.save(filename, upload_marks)
            comedk.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('comedk-ranking-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        initial_data = {'taken_comedk': 'True'}
        form = COMEDKForm(initial=initial_data)

    context = {
        'form': form,
    }
    return render(request, 'web-forms/comedk-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_comedk']), login_url='access-denied-page')
def view_comedk (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        comedk_list = COMEDK.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        comedk_list = COMEDK.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    comedk_list= all_paginator(comedk_list, page_number, list_par_page)
        
    context = {
        'comedk_ranking_list':comedk_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/comedk.html', context)
    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_comedk']), login_url='access-denied-page')
def delete_comedk (request, pk):
    comedk = get_object_or_404(COMEDK, pk=pk)
    comedk.delete()
    return redirect('comedk-ranking-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_comedk']), login_url='access-denied-page')
def download_comedk(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        comedk_list = COMEDK.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        comedk_list = COMEDK.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'comedk'
    
    headers = ['ID', 'Created At', 'Name', 'Phone No.', 'Email', 'Registration No.', 'Specialization', 'Campus', 'Taken COMEDK', 'COMEDK Rank', 'COMEDK Registration No.']
    sheet.append(headers)
    
    for item in comedk_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.registration_number,
            item.specialization,
            item.campus,
            "Yes" if item.taken_comedk else "No",
            item.comedk_rank if item.comedk_rank else "-",
            item.tat_number if item.tat_number else "-",
            item.comedk_registration_number if item.comedk_registration_number else "-",
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=comedk.xlsx'
    return response




def jee_main2_form (request):
    if request.method == 'POST':
        form = JEEMain2Form(request.POST, request.FILES)
        if form.is_valid():
            jee_main2 = form.save(commit=False)
            upload_marks = form.cleaned_data.get('upload_marks')
            phone_number = form.cleaned_data.get('phone_number')
            if upload_marks:
                filename = create_unique_filename(upload_marks)
                jee_main2.upload_marks.save(filename, upload_marks)
            jee_main2.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('jee-main2-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        initial_data = {'taken_jee_main2': 'True'}
        form = JEEMain2Form(initial=initial_data)

    context = {
        'form': form,
    }
    return render(request, 'web-forms/jee-main2-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_jeemain2']), login_url='access-denied-page')
def view_jee_main2 (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        jee_main2_list = JEEMain2.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        jee_main2_list = JEEMain2.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    jee_main2_list= all_paginator(jee_main2_list, page_number, list_par_page)
        
    context = {
        'jee_main2_list':jee_main2_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/jee-main2.html', context)
    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_jeemain2']), login_url='access-denied-page')
def delete_jee_main2 (request, pk):
    jee_main2 = get_object_or_404(JEEMain2, pk=pk)
    jee_main2.delete()
    return redirect('jee-main2-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_jeemain2']), login_url='access-denied-page')
def download_jee_main2(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        jee_main2_list = JEEMain2.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        jee_main2_list = JEEMain2.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'jee_main2'
    
    headers = ['ID', 'Created At', 'Name', 'Phone No.', 'Email', 'Registration No.', 'Specialization', 'Campus', 'Taken JEE Main 1', 'JEE Main 1 Rank', 'JEE Main 1 Registration No.']
    sheet.append(headers)
    
    for item in jee_main2_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.registration_number,
            item.specialization,
            item.campus,
            "Yes" if item.taken_jee_main2 else "No",
            item.jee_main2_rank if item.jee_main2_rank else "-",
            item.jee_main2_registration_number if item.jee_main2_registration_number else "-",
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=jee_main2.xlsx'
    return response


def jee_main1_form (request):
    if request.method == 'POST':
        form = JEEMain1Form(request.POST, request.FILES)
        if form.is_valid():
            jee_main1 = form.save(commit=False)
            upload_marks = form.cleaned_data.get('upload_marks')
            phone_number = form.cleaned_data.get('phone_number')
            if upload_marks:
                filename = create_unique_filename(upload_marks)
                jee_main1.upload_marks.save(filename, upload_marks)
            jee_main1.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('jee-main1-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        initial_data = {'taken_jee_main1': 'True'}
        form = JEEMain1Form(initial=initial_data)

    context = {
        'form': form,
    }
    return render(request, 'web-forms/jee-main1-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_jeemain1']), login_url='access-denied-page')
def view_jee_main1 (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        jee_main1_list = JEEMain1.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        jee_main1_list = JEEMain1.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    jee_main1_list= all_paginator(jee_main1_list, page_number, list_par_page)
        
    context = {
        'jee_main1_list':jee_main1_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/jee-main1.html', context)
    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_jeemain1']), login_url='access-denied-page')
def delete_jee_main1 (request, pk):
    jee_main1 = get_object_or_404(JEEMain1, pk=pk)
    jee_main1.delete()
    return redirect('jee-main1-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_jeemain1']), login_url='access-denied-page')
def download_jee_main1(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        jee_main1_list = JEEMain1.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        jee_main1_list = JEEMain1.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'jee_main1'
    
    headers = ['ID', 'Created At', 'Name', 'Phone No.', 'Email', 'Registration No.', 'Specialization', 'Campus', 'Taken JEE Main 1', 'JEE Main 1 Rank', 'JEE Main 1 Registration No.']
    sheet.append(headers)
    
    for item in jee_main1_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.registration_number,
            item.specialization,
            item.campus,
            "Yes" if item.taken_jee_main1 else "No",
            item.jee_main1_rank if item.jee_main1_rank else "-",
            item.jee_main1_registration_number if item.jee_main1_registration_number else "-",
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=jee_main1.xlsx'
    return response


def request_a_call_back (request):
    return render(request, 'base/request-call-back.html')

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['add_dailyreport']), login_url='access-denied-page')
def add_daily_report(request):
    daily_report_list = DailyReport.objects.filter(user=request.user).order_by('-date', '-created_at')[:5]
    initial_data = {'date': date.today(), 'category': 'Web Update'}
    form = DailyReportForm(initial=initial_data)
    if request.method == 'POST':
        form = DailyReportForm(request.POST)
        if form.is_valid():
            daily_report = form.save(commit=False)
            daily_report.user = request.user  # Set the user field here
            daily_report.save()
            return redirect('add-daily-report')
        else:
            return redirect('index')
    
    context={
        'form':form,
        'daily_report_list':daily_report_list,
    } 
    return render(request, 'web-forms/daily-report-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['add_dailyreport']), login_url='access-denied-page')
def view_daily_report (request, username):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        daily_report_list = DailyReport.objects.filter(user__username=username, date__range=[start_datetime, end_datetime])
            
    else:
        daily_report_list = DailyReport.objects.filter(user__username=username)
        
    page_number = request.GET.get('page')
    list_par_page = 25
    daily_report_list= all_paginator(daily_report_list, page_number, list_par_page)
        
    context = {
        'daily_report_list':daily_report_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/daily-report.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_dailyreport']), login_url='access-denied-page')
def view_all_daily_report (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert start_date and end_date to date objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Filter using date range
        daily_report_list = DailyReport.objects.filter(date__range=[start_date, end_date])
    else:
        daily_report_list = DailyReport.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    daily_report_list= all_paginator(daily_report_list, page_number, list_par_page)
        
    context = {
        'daily_report_list':daily_report_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/daily-report.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['change_dailyreport']), login_url='access-denied-page')
def edit_daily_report(request, pk):
    daily_report_list = DailyReport.objects.filter(user=request.user).order_by('-date', '-created_at')[:5]
    edit_daily_report= get_object_or_404(DailyReport, pk=pk)
    form = DailyReportForm(instance=edit_daily_report)
    if request.method == 'POST':
        form = DailyReportForm(request.POST, instance=edit_daily_report)
        if form.is_valid():
            # Associate the club with the award before saving
            form.save()
            return redirect('add-daily-report')
        else:
            return redirect('edit-daily-report', pk)
    
    context={
        'form':form,
        'daily_report_list':daily_report_list,
    } 
    return render(request, 'web-forms/daily-report-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_dailyreport']), login_url='access-denied-page')
def delete_daily_report(request, pk):
    delete_daily_report = get_object_or_404(DailyReport, pk=pk)
    delete_daily_report.delete()
    return redirect('add-daily-report')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_dailyreport']), login_url='access-denied-page')
def download_daily_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert start_date and end_date to date objects
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        # Filter using date range
        daily_report_list = DailyReport.objects.filter(date__range=[start_date, end_date])
    else:
        daily_report_list = DailyReport.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'daily_report'
    
    headers = [
        'ID', 'Date', 'Campus', 'Title', 'Description', 'Status', 'Category'
        ]
    sheet.append(headers)
    
    for item in daily_report_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            item.date,
            item.campus,
            item.title,
            item.description,
            item.status,
            item.category,
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=daily_report.xlsx'
    return response


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_permission']), login_url='access-denied-page')
def add_group_permissions(request):
    form = GroupPermissionForm()
    user_groups = Group.objects.all()
    # if check_requested_user_permissions(request.user, ['add_permission']):
    if request.method == 'POST':
        form = GroupPermissionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('create-group-permissions')
            
    context = {
        'form': form,
        'user_groups':user_groups
    }
    return render(request, 'admin/group_permissions.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['change_permission']), login_url='access-denied-page')
def edit_group_permissions(request, group_id):
    user_groups = Group.objects.all()
    group = Group.objects.get(id=group_id)
    if request.method == 'POST':
        form = GroupPermissionForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('create-group-permissions')
    else:
        form = GroupPermissionForm(instance=group)
    context = {
        'form': form,
        'group': group,
        'user_groups':user_groups
    }
    return render(request, 'admin/group_permissions.html', context)







def puc_marks_upload_form (request):
    if request.method == 'POST':
        form = PUCUpoloadMarksForm(request.POST, request.FILES)
        if form.is_valid():
            puc_marks_upload = form.save(commit=False)
            upload_marks = form.cleaned_data.get('upload_marks')
            phone_number = form.cleaned_data.get('phone_number')
            if upload_marks:
                filename = create_unique_filename(upload_marks)
                puc_marks_upload.upload_marks.save(filename, upload_marks)
            puc_marks_upload.save()
                
            messages.success(request, 'Form submitted successfully!')
            return redirect('puc-marks-upload-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        form = PUCUpoloadMarksForm()

    context = {
        'form': form,
    }
    return render(request, 'web-forms/puc-marks-upload-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_pucupoloadmarks']), login_url='access-denied-page')
def view_puc_marks_upload(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        puc_marks_upload_list = PUCUpoloadMarks.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        puc_marks_upload_list = PUCUpoloadMarks.objects.all()
    
    # storage_client = storage.Client()
    # bucket = storage_client.bucket(settings.BUCKET_NAME)
    # # Generate signed URLs for each upload and check if the file exists
    # for upload in puc_marks_upload_list:
    #     if upload.upload_marks:
    #         blob = bucket.blob(upload.upload_marks.name)
    #         if blob.exists():
    #             upload.signed_url = generate_signed_url(upload.upload_marks.name)
    #         else:
    #             upload.signed_url = None

    page_number = request.GET.get('page')
    list_par_page = 25
    puc_marks_upload_list = all_paginator(puc_marks_upload_list, page_number, list_par_page)
    
    context = {
        'puc_marks_upload_list': puc_marks_upload_list,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'users/puc-marks-upload.html', context)



# def view_puc_marks_upload (request):
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     if start_date and end_date:
#         # Convert the start and end date to datetime objects
#         start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
#         end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
#         puc_marks_upload_list = PUCUpoloadMarks.objects.filter(created_at__range=[start_datetime, end_datetime])
#     else:
#         puc_marks_upload_list = PUCUpoloadMarks.objects.all()
        
#     page_number = request.GET.get('page')
#     list_par_page = 25
#     puc_marks_upload_list= all_paginator(puc_marks_upload_list, page_number, list_par_page)
        
#     context = {
#         'puc_marks_upload_list':puc_marks_upload_list,
#         'start_date':start_date,
#         'end_date':end_date,
#     }
#     return render(request, 'users/puc-marks-upload.html', context)




    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_pucupoloadmarks']), login_url='access-denied-page')
def delete_puc_marks_upload (request, pk):
    puc_marks_upload = get_object_or_404(PUCUpoloadMarks, pk=pk)
    puc_marks_upload.delete()
    return redirect('puc-marks-upload-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_pucupoloadmarks']), login_url='access-denied-page')
def download_puc_marks_upload(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        puc_marks_upload_list = PUCUpoloadMarks.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        puc_marks_upload_list = PUCUpoloadMarks.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'puc_marks_upload'
    
    headers = [
        'ID', 'Created At', 'Name', 'Phone No.', 'Email', 'Registration No.', 'Specialization', 'Campus', 'Board', 'Intermediate Candidates',
        'Physics', 'Chemistry', 'Mathematics A', 'Mathematics B', 'Electronics', 'Computer Science', 'Biotechnology', 'Percentage', 'Comment'
        ]
    sheet.append(headers)
    
    for item in puc_marks_upload_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.registration_number,
            item.specialization,
            item.campus,
            item.board,
            "Yes" if item.intermediate_candidates else "No",
            item.physics,
            item.chemistry,
            item.mathematics_a,
            item.mathematics_b if item.mathematics_b else "-",
            item.electronics,
            item.computer_science if item.computer_science else "-",
            item.biotechnology if item.biotechnology else "-",
            item.aggregate_percentage,
            item.comment,
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=puc_marks_upload.xlsx'
    return response


def cet_ranking_form (request):
    if request.method == 'POST':
        form = CETRankingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('cet-ranking-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        initial_data = {'taken_kcet': 'True'}
        form = CETRankingForm(initial=initial_data)

    context = {
        'form': form,
    }
    return render(request, 'web-forms/cet-ranking-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_cetranking']), login_url='access-denied-page')
def view_cet_ranking (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        cet_ranking_list = CETRanking.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        cet_ranking_list = CETRanking.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    cet_ranking_list= all_paginator(cet_ranking_list, page_number, list_par_page)
        
    context = {
        'cet_ranking_list':cet_ranking_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/cet-ranking.html', context)
    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_cetranking']), login_url='access-denied-page')
def delete_cet_ranking (request, pk):
    cet_ranking = get_object_or_404(CETRanking, pk=pk)
    cet_ranking.delete()
    return redirect('cet-ranking-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_cetranking']), login_url='access-denied-page')
def download_cet_ranking(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        cet_ranking_list = CETRanking.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        cet_ranking_list = CETRanking.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'cet_ranking'
    
    headers = ['ID', 'Created At', 'Name', 'Email', 'Phone No.', 'Registration No.', 'Specialization', 'Campus', 'Taken KCET', 'KCET Rank', 'KCET Registration No.']
    sheet.append(headers)
    
    for item in cet_ranking_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.email,
            item.phone_number,
            item.registration_number,
            item.specialization,
            item.campus,
            "Yes" if item.taken_kcet else "No",  # Convert Boolean to "Yes"/"No"
            item.kcet_rank if item.kcet_rank else "-",  # Fallback to "-"
            item.kcet_registration_number if item.kcet_registration_number else "-"  # Fallback to "-"
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cet_ranking.xlsx'
    return response


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['add_faq']), login_url='access-denied-page')
def faq_form(request):
    faq_list = FAQ.objects.all()[:5]
    form = FAQForm()
    if request.method == 'POST':
        form = FAQForm(request.POST)
        if form.is_valid():
            # Associate the club with the award before saving
            faq = form.save(commit=False)
            faq.save()
            # category = form.cleaned_data['category']
            # faq.category.set(category)
            return redirect('add-faq')
        else:
            return redirect('add-faq')
    
    context={
        'form':form,
        'faq_list':faq_list,
    }
    return render(request, 'web-forms/faq-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_faq']), login_url='access-denied-page')
def view_faq (request): 
    q = request.GET.get('search') if request.GET.get('search') != None else ''   
    # Fetch all FAQs
    all_faqs = FAQ.objects.all()
    faq_list = []

    if q:
        for faq in all_faqs:
            if fuzz.partial_ratio(q.lower(), faq.question.lower()) > 70 or \
                fuzz.partial_ratio(q.lower(), faq.answer.lower()) > 70 or \
                fuzz.partial_ratio(q.lower(), faq.category.lower()) > 70:
                faq_list.append(faq)
        faq_list = sorted(faq_list, key=lambda faq: fuzz.partial_ratio(q.lower(), faq.question.lower()), reverse=True)
    else:
        faq_list = FAQ.objects.filter(Q(question__icontains=q) | Q(answer__icontains=q) | Q(category__icontains=q)).order_by('category')


    page_number = request.GET.get('page')
    list_par_page = 25
    faq_list= all_paginator(faq_list, page_number, list_par_page)
        
    context = {
        'faq_list':faq_list,
    }
    return render(request, 'users/faq.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['change_faq']), login_url='access-denied-page')
def edit_faq(request, pk):
    faq_list = FAQ.objects.all()
    edit_faq = faq_list.get(pk=pk)
    form = FAQForm(instance=edit_faq)
    if request.method == 'POST':
        form = FAQForm(request.POST, instance=edit_faq)
        if form.is_valid():
            # Associate the club with the award before saving
            faq = form.save(commit=False)
            faq.save()
            # category = form.cleaned_data['category']
            # faq.category.set(category)
            return redirect('add-faq')
        else:
            return redirect('add-faq')
    
    context={
        'form':form,
        'faq_list':faq_list[:5],
    }
    return render(request, 'web-forms/faq-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_faq']), login_url='access-denied-page')
def delete_faq(request, pk):
    delete_faq = get_object_or_404(FAQ, pk=pk)
    delete_faq.delete()
    return redirect('add-faq')


# Create your views here.
def index (request):
    return render(request, 'base/index.html')


def category(request, category_name=None):
    q = request.GET.get('search') if request.GET.get('search') != None else ''
    if category_name:
        faq_list = FAQ.objects.filter(category__icontains=category_name)
    else:
        # Fetch all FAQs
        all_faqs = FAQ.objects.all()
        faq_list = []

        if q:
            for faq in all_faqs:
                if fuzz.partial_ratio(q.lower(), faq.question.lower()) > 70 or \
                   fuzz.partial_ratio(q.lower(), faq.answer.lower()) > 70 or \
                   fuzz.partial_ratio(q.lower(), faq.category.lower()) > 70:
                    faq_list.append(faq)
            faq_list = sorted(faq_list, key=lambda faq: fuzz.partial_ratio(q.lower(), faq.question.lower()), reverse=True)
        else:
            faq_list = FAQ.objects.filter(Q(question__icontains=q) | Q(answer__icontains=q) | Q(category__icontains=q)).order_by('category')

        category_name = q

    context = {
        'category_name': category_name,
        'faq_list': faq_list,
    }
    return render(request, 'base/category.html', context)

def search_category(request, faq_table_name, category_name):
    q = request.GET.get('search') if request.GET.get('search') != None else ''
    if category_name:
        faq_list = faq_table_name.objects.filter(category__category__icontains=category_name)
    else:
        # Fetch all FAQs
        all_faqs = faq_table_name.objects.all()
        faq_list = []

        if q:
            for faq in all_faqs:
                if fuzz.partial_ratio(q.lower(), faq.question.lower()) > 70 or \
                fuzz.partial_ratio(q.lower(), faq.answer.lower()) > 70 or \
                fuzz.partial_ratio(q.lower(), faq.category.category.lower()) > 70:  # Access the category name
                    faq_list.append(faq)
            faq_list = sorted(faq_list, key=lambda faq: fuzz.partial_ratio(q.lower(), faq.question.lower()), reverse=True)
        else:
            faq_list = faq_table_name.objects.filter(Q(question__icontains=q) | Q(answer__icontains=q) | Q(category__icontains=q)).order_by('category')

        category_name = q
    return faq_list


def selected_category(request, model_name, category_name):
    model_class = model_mapping.get(model_name)
    category = category_model_mapping.get(model_name)
    if not model_class or not category:
        return render(request, 'error/404.html')
    
    category_list = category.objects.all()
    faq_list = search_category(request, model_class, category_name)
    context = {
        'model_name':model_name,
        'category_list': category_list,
        'category_name': category_name,
        'faq_list': faq_list,
    }
    return render(request, 'base/category.html', context)






@login_required(login_url='login-page')
@user_passes_test(lambda u: u.is_active, login_url='access-denied-page')
def dashboard (request):
    
    context = {
    }
    return render(request, 'users/dashboard.html', context)


# login and logout 
def loginpage(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        remember = request.POST.get('remember_me') == 'on'  # Check if the checkbox is selected

        user = authenticate(request, username=username, password=password)

        if user is not None:
            if remember:
                request.session.set_expiry(settings.SESSION_COOKIE_AGE_REMEMBER_ME)  # Set longer session cookie age
            else:
                request.session.set_expiry(settings.SESSION_COOKIE_AGE)  # Set default session cookie age
                
            if user.is_superuser or user.groups.exists():
                login(request, user)
                return redirect('dashboard')
            
            if user != user.is_superuser:
                login(request, user)
                return redirect('index')
            
        else:
            # Check if the username exists
            if User.objects.filter(username=username).exists():
                messages.error(request, "The password you entered is incorrect.")
            else:
                messages.error(request, "The username you entered does not exist.")

    return render(request, 'login_users/login-form.html')


@login_required(login_url='login-page')
def logoutpage(request):
    logout(request)
    return redirect('index')


from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.views import PasswordChangeView
from .login_form import PasswordChangingForm
from django.urls import reverse_lazy
class PasswordsChangeView(PasswordChangeView):
    form_class = PasswordChangingForm
    success_url = reverse_lazy('index')
    
    
@login_required(login_url='login-page')
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangingForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            # messages.success(request, 'Your password has been successfully updated')
            return redirect('login-page')
        else:
            messages.error(request, 'Your password should include a mix of uppercase and lowercase letters, numbers, and symbols.')
    else:
        form = PasswordChangingForm(request.user)
    context= {'form': form}
    return render(request, 'login_users/change_password.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['change_user',]), login_url='access-denied-page')
def admin_changing_users_passwords(request, user_id):
    get_user= get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        password1=request.POST['password1']
        password2=request.POST['password2']
        if password1 == password2:
            get_user.password= make_password(password1)
            get_user.save()  # save the updated password to the database 
            return redirect('dashboard')
        else:
            messages.success(request, "Passwords you entered don't match. Please try again.")
            
    context={ 'get_user':get_user}
    return render(request, 'login_users/admin_change_password.html', context)



def all_paginator(news_obj, page_number, news_par_page):
    p = Paginator(news_obj, news_par_page) 
    try:
        news_obj = p.get_page(page_number) 
    except PageNotAnInteger:
        news_obj = p.page(1)
    except EmptyPage:
        news_obj = p.page(p.num_pages)
    return news_obj


def grievance_form (request):
    if request.method == 'POST':
        form = GrievanceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('grievance-form')
        else:
            messages.success(request, 'Form is invalide..')
    else:
        form = GrievanceForm()
        
    context= {
        'form':form,
    }
    return render(request, 'web-forms/grievance-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_grievance']), login_url='access-denied-page')
def view_grievance (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        grievance_list = Grievance.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        grievance_list = Grievance.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    grievance_list= all_paginator(grievance_list, page_number, list_par_page)
        
    context = {
        'grievance_list':grievance_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/grievance.html', context)
    

@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_grievance']), login_url='access-denied-page')
def delete_grievance (request, pk):
    grievance = get_object_or_404(Grievance, pk=pk)
    grievance.delete()
    return redirect('grievance-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_grievance']), login_url='access-denied-page')
def download_grievance(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        grievance_list = Grievance.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        grievance_list = Grievance.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Grievance'
    
    headers = ['Id', 'Created At', 'Name', 'Phone Number', 'Email', 'Designation', 'Identity', 'Identity Number', 'Types of Grievance', 'Your Grievance']
    sheet.append(headers)
    
    for item in grievance_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.designation,
            item.identity,
            item.identity_number,
            item.types_of_grievance,
            item.your_grievance
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Grievance.xlsx'
    return response


def amaatra_form (request):
    if request.method == 'POST':
        form = AmaatraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Form submitted successfully!')
            return redirect('amaatra-form')
        else:
            messages.success(request, 'Form is invalide..')
    else:
        form = AmaatraForm()
        
    context= {
        'form':form,
    }
    return render(request, 'web-forms/amaatra-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_amaatra']), login_url='access-denied-page')
def view_amaatra (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # Filter the queryset using the datetime range
        amaatra_list = Amaatra.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        amaatra_list = Amaatra.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    amaatra_list= all_paginator(amaatra_list, page_number, list_par_page)
    
    context = {
        'amaatra_list':amaatra_list,
        'start_date':start_date,
        'end_date':end_date,
        
    }
    return render(request, 'users/amaatra.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_amaatra']), login_url='access-denied-page')
def delete_amaatra (request, pk):
    amaatra = get_object_or_404(Amaatra, pk=pk)
    amaatra.delete()
    return redirect('amaatra-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_amaatra']), login_url='access-denied-page')
def download_amaatra(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # Filter the queryset using the datetime range
        amaatra_list = Amaatra.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        amaatra_list = Amaatra.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Amaatra'
    
    headers = ['ID', 'Created At', 'Name', 'Phone Number', 'Email', 'Address']
    sheet.append(headers)
    
    for item in amaatra_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.address
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Amaatra.xlsx'
    return response


def ssm_form (request):
    if request.method == 'POST':
        form = SSMForm(request.POST)
        if form.is_valid():
            ssm = form.save(commit=False)
            messages.success(request, 'Form submitted successfully!')
            ssm.save()
            return redirect('ssm-form')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
    else:
        form = SSMForm()
        
    context= {
        'form':form,
    }
    return render(request, 'web-forms/ssm-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_ssm']), login_url='access-denied-page')
def view_ssm (request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # Filter the queryset using the datetime range
        ssm_list = SSM.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        ssm_list = SSM.objects.all()
        
    page_number = request.GET.get('page')
    list_par_page = 25
    ssm_list= all_paginator(ssm_list, page_number, list_par_page)
    
    context = {
        'ssm_list':ssm_list,
        'start_date':start_date,
        'end_date':end_date,
    }
    return render(request, 'users/ssm.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_ssm']), login_url='access-denied-page')
def delete_ssm (request, pk):
    ssm = get_object_or_404(SSM, pk=pk)
    ssm.delete()
    return redirect('ssm-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_ssm']), login_url='access-denied-page')
def download_ssm(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        # Convert the start and end date to datetime objects
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # Filter the queryset using the datetime range
        ssm_list = SSM.objects.filter(created_at__range=[start_datetime, end_datetime])
    else:
        ssm_list = SSM.objects.all()
    
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'EC Transportation'
    
    headers = ['ID', 'Created At', 'Name', 'Phone Number', 'Email', 'Class to which admission is sought']
    sheet.append(headers)
    
    for item in ssm_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            created_at_naive,
            item.name,
            item.phone_number,
            item.email,
            item.class_admission
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=SSM.xlsx'
    return response


def ec_transportation_form (request):
    if request.method == 'POST':
        form = ECTransportationForm(request.POST, request.FILES, route=request.POST.get('route'))
        if form.is_valid():
            ec_transportation = form.save(commit=False)
            photo = form.cleaned_data.get('photo')
            phone_number = form.cleaned_data.get('phone_number')
            if photo:
                filename = create_unique_filename(photo)
                ec_transportation.photo.save(filename, photo)
            ec_transportation.save()
            messages.success(request, "Form submitted successfully.")
            return redirect('ec-transportation-form')
    else:
        form = ECTransportationForm()
    context = {
        'form': form,
    }
    return render(request, 'web-forms/ec-transportation-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_ectransportation']), login_url='access-denied-page')
def delete_ec_transportation (request, pk):
    ec_transportation = get_object_or_404(ECTransportation, pk=pk)
    ec_transportation.delete()
    return redirect('ec-transportation-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_ectransportation']), login_url='access-denied-page')
def view_ec_transportation(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # ec_transportation_list = ECTransportation.objects.filter(created_at__range=[start_datetime, end_datetime])
        ec_transportation_list = ECTransportation.objects.filter(dob__range=[start_datetime, end_datetime])
    else:
        ec_transportation_list = ECTransportation.objects.all()
    
    page_number = request.GET.get('page')
    list_par_page = 25
    ec_transportation_list= all_paginator(ec_transportation_list, page_number, list_par_page)

    context = {
        'ec_transportation_list': ec_transportation_list,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'users/ec_transportation.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_ectransportation']), login_url='access-denied-page')
def download_ec_transportation(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    # ec_transportation_list = ECTransportation.objects.filter(created_at__range=[start_date, end_date])
    if start_date and end_date:
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        # ec_transportation_list = ECTransportation.objects.filter(created_at__range=[start_datetime, end_datetime])
        ec_transportation_list = ECTransportation.objects.filter(dob__range=[start_datetime, end_datetime])
    else:
        ec_transportation_list = ECTransportation.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'EC Transportation'
    
    headers = ['ID', 'Name', 'Phone Number', 'Parent Phone Number', 'Email', 'Address', 'DOB', 'Blood Group', 'Registration No', 'Program', 'Route', 'Pickup Point', 'Created At', 'Marks Sheet']
    sheet.append(headers)
    
    for item in ec_transportation_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            item.name,
            item.phone_number,
            item.parent_phone_number,
            item.email,
            item.address,
            item.dob,
            item.blood_group,
            item.registration_no,
            item.program,
            item.route,
            item.pickup_point,
            created_at_naive,
            str(f'{item.pk}_{item.photo.name.split("/")[-1]}') if item.photo else "-",
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EC_Transportation.xlsx'
    return response



def rr_transportation_form(request):
    if request.method == 'POST':
        form = RRTransportationForm(request.POST, request.FILES, route=request.POST.get('route'))
        
        if form.is_valid():
            rr_transportation = form.save(commit=False)
            photo = form.cleaned_data.get('photo')
            phone_number = form.cleaned_data.get('phone_number')
            
            if photo:
                filename = create_unique_filename(photo)
                rr_transportation.photo.save(filename, photo)
            rr_transportation.save()
            messages.success(request, "Form submitted successfully.")
            return redirect('rr-transportation-form')
        else:
            messages.error(request, "There was an error submitting the form.")
    else:
        form = RRTransportationForm()
        
    context = {
        'form': form,
    }
    return render(request, 'web-forms/rr-transportation-form.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['delete_rrtransportation']), login_url='access-denied-page')
def delete_rr_transportation (request, pk):
    rr_transportation = get_object_or_404(RRTransportation, pk=pk)
    rr_transportation.delete()
    return redirect('rr-transportation-list')


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_rrtransportation']), login_url='access-denied-page')
def view_rr_transportation(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        rr_transportation_list = RRTransportation.objects.filter(dob__range=[start_datetime, end_datetime])
    else:
        rr_transportation_list = RRTransportation.objects.all()
    
    page_number = request.GET.get('page')
    list_par_page = 25
    rr_transportation_list= all_paginator(rr_transportation_list, page_number, list_par_page)

    context = {
        'rr_transportation_list': rr_transportation_list,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'users/rr_transportation.html', context)


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_rrtransportation']), login_url='access-denied-page')
def download_rr_transportation(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time(23, 59, 59))
        rr_transportation_list = RRTransportation.objects.filter(dob__range=[start_datetime, end_datetime])
    else:
        rr_transportation_list = RRTransportation.objects.all()
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'EC Transportation'
    
    headers = ['ID', 'Name', 'Phone Number', 'Parent Phone Number', 'Email', 'Address', 'DOB', 'Blood Group', 'Registration No', 'Program', 'Route', 'Pickup Point', 'Created At', 'Marks Sheet']
    sheet.append(headers)
    
    for item in rr_transportation_list:
        created_at_naive = item.created_at.replace(tzinfo=None)  # Remove timezone info
        sheet.append([
            item.pk,
            item.name,
            item.phone_number,
            item.parent_phone_number,
            item.email,
            item.address,
            item.dob,
            item.blood_group,
            item.registration_no,
            item.program,
            item.route,
            item.pickup_point,
            created_at_naive,
            str(f'{item.pk}_{item.photo.name.split("/")[-1]}') if item.photo else "-",
        ])
    
    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=rr_Transportation.xlsx'
    return response


@login_required(login_url='login-page')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['view_group']), login_url='access-denied-page')
def create_admin_user(request):
    form = AdminUserCreationForm()
    # In database we have user_groups table and access that table we need this
    if request.user.is_superuser:
        user_groups = User.groups.through.objects.select_related('user', 'group')
    else:
        user_groups = User.groups.through.objects.select_related('user', 'group').filter(user__is_superuser=False)
    unique_users = set()
    for user_group in user_groups:
        unique_users.add(user_group.user)
        
    # if check_requested_user_permissions(request.user, ['add_group']):
    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            # admin_user = form.save(commit=False)
            admin_user = form.save()
            groups = form.cleaned_data['groups']
            admin_user.groups.set(groups)
            
            messages.success(request, 'User created successfully.')
            return redirect('create-admin-user')  # Replace 'user-list' with the URL name of the user list view
     
    context={
        'form': form,
        'user_groups': unique_users,
    }
    
    return render(request, 'admin/create_admin_user.html', context)


@login_required(login_url='login-users-form')
@user_passes_test(lambda user: check_requested_user_permissions(user, ['change_group']), login_url='access-denied-page')
def edit_admin_user(request, username):
    if request.user.is_superuser:
        user_groups = User.groups.through.objects.select_related('user', 'group')
    else:
        user_groups = User.groups.through.objects.select_related('user', 'group').filter(user__is_superuser=False)
    unique_users = set()
    for user_group in user_groups:
        unique_users.add(user_group.user)
        
    groups = Group.objects.prefetch_related('user_set').all()
    admin_user = User.objects.get(username=username) 
    if request.method == 'POST':
        form = AdminUserEditForm(request.POST, request.FILES, instance=admin_user)
        if form.is_valid():
            admin_user = form.save()
            groups = form.cleaned_data['groups']
            admin_user.groups.set(groups)
            messages.success(request, 'user updated successfully.')
            return redirect('create-admin-user')  
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label
                # Display only the first error for the first field with an error
                first_error = errors[0]  # Get the first error message
                messages.error(request, f"Error in {label}: {first_error}")
                break  # Stop after the first field with an error
            
    else:
        form = AdminUserEditForm(instance=admin_user)

    context={
        'form': form,
        'user_groups':unique_users,
        'groups':groups,
    }
    return render(request, 'admin/create_admin_user.html', context)


@login_required(login_url='login-page')
# @user_passes_test(lambda user: check_requested_user_permissions(user, ['view_user']), login_url='access-denied-page')
def profile(request, username):
    user_profile = get_object_or_404(User, username=username)
    
    context={
        'user_profile':user_profile
    }
    return render(request, 'users/profile.html', context)

def contact_us(request):
    return render(request, 'base/contactus.html')

# to handel error pages 
def error_404_view(request, exception=None):
    return render(request, 'error/404.html')

# to handel access denied pages 
def access_denied_page(request):
    return render(request, 'error/access_denied_page.html')